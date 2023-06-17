import shutil

from flask import (
    Blueprint,
    render_template,
    redirect,
    request,
    flash,
    make_response,
    send_from_directory,
)
import xlrd
import openpyxl
import datetime
from werkzeug.utils import secure_filename
import os
from pathlib import Path
from flask_login import login_required

# BASE_DIR = "/www/test4.barsinfo.ru/uploads_files/"
UPLOAD_FILE = Path("upload_files")
DOWNLOAD_DIR = Path("download_files")
# DOWNLOAD_DIR = "/www/test4.barsinfo.ru/download_files/"


input_file = Blueprint("input_file", __name__)

ALLOWED_EXTENSIONS = {"xls", "xlsx"}


@input_file.route("/", methods=["POST"])
@login_required
def get_file():
    file = request.files["file"]

    if file.filename == "":
        flash("Нет выбранного файла")
        return redirect(request.url)

    if file:
        # безопасно извлекаем оригинальное имя файла
        filename = secure_filename(file.filename)
        # сохраняем файл
        file.save(os.path.join(UPLOAD_FILE, filename))

        p = []

        if filename.endswith(".xls"):
            wb = xlrd.open_workbook(file)
            sheet = wb.sheet_by_index(0)
            row = []

            for i in range(
                sheet.ncols
            ):  # запускаем цикл перебора данных до значения максимального количества столбцов
                if (
                    sheet.cell(2, i).ctype
                ) != 3:  # если значение ячейки не дата, то записываем его значение в список (3 - значение типа данных даты)
                    row.append(sheet.cell_value(2, i))
                else:
                    cell = sheet.cell_value(2, i)
                    date = xlrd.xldate.xldate_as_datetime(cell, wb.datemode)
                    date = date.strftime("%d.%m.%Y")
                    row.append(date)

            p = row
        elif filename.endswith(".xlsx"):
            wb = openpyxl.load_workbook(file)
            sheet = wb.active
            values = [
                sheet.cell(row=3, column=i).value
                for i in range(1, sheet.max_column + 1)
            ]
            values_date = []
            for i in values:
                if isinstance(i, datetime.datetime):
                    values_date.append(datetime.datetime.strftime(i, "%d.%m.%Y"))
                else:
                    values_date.append(i)
            p = values_date

        fk_vtd = ["FK", "TXVT170101", "АСФК", "22.0", ""]
        from_vtd = [
            "FROM",
            "1100",
            "Управление Федерального казначейства по Республике Татарстан",
        ]
        to_vtd = ["TO", "2", "92200111", "МИНИСТЕРСТВО ФИНАНСОВ РЕСПУБЛИКИ ТАТАРСТАН"]
        vt_vtd = [
            "VT",
            "F20D9208-383B-047C-E053-0A0B052CA007",
            "04112001110",
            "11.01.2023",
            "10.01.2023",
            "0",
            "1100",
            "Управление Федерального казначейства по Республике Татарстан",
            "92200111",
            "Министерство финансов Республики Татарстан",
            "711",
            "Министерство финансов Республики Татарстан",
            "бюджет Республики Татарстан",
            "92000000",
            "02301384",
            "Министерство финансов Республики Татарстан",
            "Главный казначей",
            "Хусаинова А.М.",
            "8(843) 528-5323",
            "12.01.2023",
            "5993811.96",
            "0.00",
            "0.00",
            "0.00",
            "0.00",
        ]
        vtsum_vtd = [
            "VTSUM",
            "91930932.79",
            "0.00",
            "0.00",
            "-3665084.62",
            "0.00",
            "97924744.75",
            "0.00",
            "0.00",
            "-3665084.62",
            "0.00",
        ]
        vtoper_vtd = [
            "VTOPER",
            "F1FC73EF-6F9B-05CE-E053-0A0B052C221A",
            "PL",
            "1",
            "11.01.2023",
            "",
            "",
            "",
            "54100.00",
            "0.00",
            "0.00",
            "",
            "20",
            "71120220086020000150",
            "",
            "92701000",
            "1654019570",
            "165501001",
            "  ",
        ]

        fk_bdd = ["FK", "TXBD230102", "ППО АСФК", "", ""]
        from_bdd = [
            "FROM",
            "1100",
            "Управление Федерального казначейства по Республике Татарстан",
        ]
        to_bdd = [
            "TO",
            "",
            "",
            "2",
            "92200111",
            "МИНИСТЕРСТВО ФИНАНСОВ РЕСПУБЛИКИ ТАТАРСТАН",
        ]
        bd_bdd = [
            "BD",
            "04112001133",
            "11.01.2023",
            "F20D9208-383B-047C-E053-0A0B052CA007",
            "VT",
            "0",
            "11",
            "5993811.96",
        ]
        bdpd_bdd = [
            "BDPD",
            "449113",
            "11.01.2023",
            "9205805000",
            "1",
            "11.01.2023",
            "4809.50",
            "0",
            "11.01.2023",
            "11.01.2023",
            "01",
            "162301112880",
            "0",
            "ИП Мунирова Гульсина Бариевна",
            "40802810000240000453",
            "049205805",
            'ПАО "АК БАРС" БАНК г. Казань',
            "30101810000000000805",
            "1626007212",
            "162601001",
            'Министерство финансов РТ (ГКУ "Социальный приют для детей и подростков "Надежда"" л/c ЛР267160006-СПНадежд)',
            "03221643920000001100",
            "019205400",
            "ОТДЕЛЕНИЕ-НБ РЕСПУБЛИКА ТАТАРСТАН БАНКА РОССИИ//УФК по Республике Татарстан, г Казань",
            "40102810445370000079",
            "",
            "",
            "5",
            "0",
            "",
            "л/c ЛР267160006-СПНадежд .Обеспечение исполнения контракта  по результатам  аукциона 0311200032622000007 НДС не облагается",
            "08",
            "0",
            "0",
            "0",
            "0",
            "0",
            "0",
            "0",
            "",
            "",
            "",
            "",
            "",
            "",
            "11.01.2023",
            "F1FC73EF-6F9B-05CE-E053-0A0B052C221A",
            "",
            "",
            "",
        ]
        bdpdst_bdd = [
            "BDPDST",
            "71111701020020000180",
            "20",
            "",
            "",
            "92000000",
            "4809.50",
            "",
            "0",
            "",
            "",
            "",
        ]

        bd_bdd[1] = p[9]
        bd_bdd[2] = p[2]
        bd_bdd[3] = p[1]
        bd_bdd[7] = p[5]

        bdpd_bdd[1] = p[0]
        bdpd_bdd[2] = p[2]
        bdpd_bdd[4] = p[0]
        bdpd_bdd[5] = p[2]
        bdpd_bdd[6] = p[5]
        bdpd_bdd[8] = p[2]
        bdpd_bdd[9] = p[2]
        bdpd_bdd[11] = p[7]
        bdpd_bdd[12] = p[6]
        bdpd_bdd[13] = p[8]
        bdpd_bdd[14] = p[10]
        bdpd_bdd[15] = p[12]
        bdpd_bdd[16] = p[13]
        bdpd_bdd[20] = p[14]
        bdpd_bdd[24] = p[15]
        bdpd_bdd[30] = p[16]
        bdpd_bdd[45] = p[2]
        bdpd_bdd[46] = p[1]

        bdpdst_bdd[3] = p[25]
        bdpdst_bdd[1] = p[24]
        bdpdst_bdd[5] = p[19]
        bdpdst_bdd[6] = p[5]

        vt_vtd[1] = p[1]
        vt_vtd[2] = p[9]
        vt_vtd[3] = p[2]  # .strftime('%d.%m.%Y')
        vt_vtd[4] = p[3]  # .strftime('%d.%m.%Y')
        vt_vtd[13] = p[19]
        vt_vtd[19] = p[2]  # .strftime('%d.%m.%Y')
        vtoper_vtd[1] = p[1]
        vtoper_vtd[3] = p[0]
        vtoper_vtd[4] = p[2]  # .strftime('%d.%m.%Y')
        vtoper_vtd[8] = p[5]
        vtoper_vtd[13] = p[21]
        vtoper_vtd[15] = p[19]

        with open(file="file_bdd.BDD", mode="w") as bdd:
            bdd.write(
                str("|".join(map(str, fk_bdd)))
                + "|"
                + "\n"
                + "|".join(map(str, from_bdd))
                + "|"
                + "\n"
                + str("|".join(map(str, to_bdd)))
                + "|"
                + "\n"
                + str("|".join(map(str, bd_bdd)))
                + "|"
                + "\n"
                + str("|".join(map(str, bdpd_bdd)))
                + "|"
                + "\n"
                + str("|".join(map(str, bdpdst_bdd)))
            )
            shutil.move(f"{bdd.name}", f"{DOWNLOAD_DIR}/{bdd.name}")

        with open(file="file_vtd.VTD", mode="w") as vtd:
            vtd.write(
                str("|".join(map(str, fk_vtd)))
                + "|"
                + "\n"
                + "|".join(map(str, from_vtd))
                + "|"
                + "\n"
                + str("|".join(map(str, to_vtd)))
                + "|"
                + "\n"
                + str("|".join(map(str, vt_vtd)))
                + "|"
                + "\n"
                + str("|".join(map(str, vtsum_vtd)))
                + "|"
                + "\n"
                + str("|".join(map(str, vtoper_vtd)))
            )
            shutil.move(f"{vtd.name}", f"{DOWNLOAD_DIR}/{vtd.name}")

        return render_template("download.html", vtd_name=vtd.name, bdd_name=bdd.name)


@input_file.route("/<path:filename>", methods=["GET"], endpoint="download")
def get_file_server(filename):
    directory_path = DOWNLOAD_DIR
    print(filename)
    path = os.path.join(DOWNLOAD_DIR, filename)
    print(path)
    # Appending app path to upload folder path within app root folder
    uploads = os.path.join(DOWNLOAD_DIR)
    # Returning file from appended path
    print(uploads)
    response = make_response(
        send_from_directory(directory_path, filename, as_attachment=True)
    )
    response.headers[
        "Content-Disposition"
    ] = f"attachment; filename={filename}"  # Задаем заголовок
    return response
